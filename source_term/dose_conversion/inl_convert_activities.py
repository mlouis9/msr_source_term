# Add python modules directory to path
import sys, os
sys.path.append(os.path.join(os.path.dirname(__file__), '../..', 'modules'))

from input_templating import lowercase_isotope_name
import pandas as pd
from collections import defaultdict
import argparse
import subprocess
from openpyxl import load_workbook
import re

seconds2years = 3.17098e-8

def add_dash_to_isotope(isotope):
    # Use a regular expression to match the element name and the atomic mass number
    match = re.match(r"([A-Za-z]+)(\d+[A-Za-z]*)", isotope)
    if match:
        element_name = match.group(1)
        a_number = match.group(2)
        return f"{element_name}-{a_number}"
    else:
        return isotope  # Return the original string if it doesn't match the pattern

def neshap_isotope_name(isotope_name):
    isotope_name = add_dash_to_isotope(lowercase_isotope_name(isotope_name))
    if isotope_name[-1] == 'M':
        isotope_name = isotope_name[:-1] + 'm'
    return isotope_name

def read_spreadsheet(input_file, input_dict={}, output_file='results.csv'):
    # Read the Excel file using pandas
    data_only = not bool(input_dict)
    workbook = load_workbook(input_file, data_only=data_only)
    sheet_name = 'Source Term Template'
    sheet = workbook[sheet_name]

    # Read the headers from the specified rows
    header_row_7 = [cell.value for cell in sheet[8] if cell.value is not None]  # Adjust the index if your header row is different

    # Combine headers to create multi-level headers
    headers = [ header for header in header_row_7 ]

    # Create a dictionary to store the data
    data = {header: [] for header in headers}

    # Populate the dictionary with cell values
    for row in sheet.iter_rows(min_row=9, values_only=True):  # Adjust the min_row if your data starts at a different row
        for header, cell_value in zip(headers, row):
            data[header].append(cell_value)

    # Access the values using headers and row indices
    nuclide_header = 'Nuclide'
    amount_used_header = 'Amount Used (Ci/yr)'
    temp_header = 'Individual Operating Temp1'

    # Find the index of the nuclide
    nuclides = [nuclide for nuclide in data[nuclide_header] if nuclide is not None]

    # Update the dictionary with the input dictionary values
    if input_dict:
        for nuclide in nuclides:
            activity_per_time, temperature = input_dict[nuclide]
            # print(activity_per_time)
            index = data[nuclide_header].index(nuclide)
            data[amount_used_header][index] = activity_per_time
            data[temp_header][index] = temperature

        # Save changes back to the workbook
        for header in headers:
            for i, value in enumerate(data[header], start=9):  # Adjust the start row if your data starts at a different row
                sheet.cell(row=i, column=headers.index(header) + 1, value=value)

        workbook.save(input_file)
    else:
        output_columns = ['Nuclide', 'Potential Effective Dose Equivalent (mrem/yr)', ' Effective Dose Equivalent (mrem/yr)', 'Percent of Unabated Dose Total']
        data = pd.DataFrame.from_dict(data)
        data = data[output_columns]
        data.to_csv(output_file, index=False)
        try:
            print(f"Total PEDE = {sheet['E6'].value:1.3E} mrem/yr\n",
                    f"Total EDE = {sheet['G6'].value:1.3E} mrem/yr\n",
                    f"PEDE % of 0.1 mrem =  {sheet['I6'].value:5f} %")
        except TypeError as e:
            print(f"Got an error when trying to print output: {e}\nTry opening up the source term spreadsheet in Excel before processing results.")


parser = argparse.ArgumentParser(
    prog='inl_convert_activities',
    description='Utility script for converting activities to effective dose using INL NESHAP spreadsheet',
)
parser.add_argument('-er', '--extract_results', action='store_true')
help_text = 'Please pass either \'mitigated\' or \'unmitigated\''
parser.add_argument('-an', '--add_nuclides', type=str, help=help_text)

# Parse the arguments
args = parser.parse_args()
operating_temp = 626.85 # °C

input_file = 'INL Source Term Template for NESHAP Calculation.xlsx'
if args.add_nuclides:
    if args.add_nuclides not in ['mitigated', 'unmitigated']:
        raise ValueError(f"The supplied --add_nuclides argument of {args.add_nuclides} is not one of the allowed types"
                         f"{help_text}")

    calc_type = (args.add_nuclides == 'mitigated')*'mitigated' + (args.add_nuclides == 'unmitigated')*'unmitigated'
    # subprocess.run(f'python calculate_{calc_type}_activities.py', shell=True)
    df = pd.read_csv(f'off_gas_activities_{calc_type}.csv')
    activities_dict = df.to_dict(orient='list') # In Ci

    # Convert activities into Ci/y by dividing by operational lifetime
    power_history = pd.read_csv('../power_history.csv').to_numpy()[:, :]
    op_lifetime = power_history[-1,0] * seconds2years
    default_func = lambda: (0.0, 0.0)
    activities_dict = defaultdict(default_func, { neshap_isotope_name(nuclide): (activity[0]/op_lifetime, operating_temp) for nuclide, activity in activities_dict.items() })

    read_spreadsheet(input_file, activities_dict)
elif args.extract_results:
    read_spreadsheet(input_file)